from collections import Counter
from statistics import mean

import numpy
import openpyxl
import pandas
from openpyxl.styles import Alignment
from sklearn.externals import joblib

# from hwgen.HWGobbler2 import make_gb_question_map
from hwgen.common import get_q_names
# from hwgen.deep.TrainTestBook import pid_override, all_page_ids, cat_page_lookup
from hwgen.deep.ttb_utils import augment_data
from hwgen.profiler import get_attempts_from_db

# gb_qmap = make_gb_question_map()


def save_class_report_card(ts, aid, gr_id, S, X, U, A, y, m_list, y_preds, slist, q_names_df, all_qids, all_page_ids=None, pid_override=None):

    months_on_list = [] 
    N = len(y_preds)
    print("c\len",N)

    print("S is",S)
    for s in S:
        months_on = s[1]/30.44
        print("appending months: ", months_on)
        months_on_list.append(months_on)
        print("MOL",months_on_list)
    months_av = mean(months_on_list)
       
    
    sum_preds = numpy.sum(y_preds, axis=0)
    print("sum of sums", numpy.sum(sum_preds))
    sum_preds = sum_preds / N
    max_sum_ix = sum_preds.argmax()
    max_sum_prob = sum_preds.max()

    vote_ct = Counter()
    for yp in y_preds:
        yp_max_ix = numpy.argmax(yp)
        label = pid_override[yp_max_ix]
        vote_ct[label]+=1

    max_vote_lab = vote_ct.most_common(1)[0][0]
    max_sum_lab = pid_override[max_sum_ix]
    print("max sum lab =", max_sum_lab, max_sum_prob)
    print("votes counted:",vote_ct.most_common(5))
    print("most voted =", max_vote_lab)


    wb = openpyxl.Workbook()
    ws = wb.active

    fn_ai = str(aid)
    r = 1
    col_headers = ["student", "age", "months_on_isaac", "qns_tried", "successes", "prev_assignts", "hexes_attempted", "top_10_topics", "last_10_qns",
                   "ISAAC_SUGGESTS", "DIFF: (too easy 1..5 too hard)", "TOPIC:(bad 1..3 good)"]
    col_widths=[len(ch) for ch in col_headers]
    for c,cv in enumerate(col_headers):
        ws.cell(r,1+c,cv)

    ws.cell(2,2, ts)

    ws.cell(2,1, "Classroom sugg'n 1:")
    ws.cell(3,1, "Classroom sugg'n 2:")
    ws.cell(2,10, max_sum_lab)
    ws.cell(3,10, max_vote_lab)

    r=4
    for s, x, u, a, t, psi, _, y_predlist in zip(S, X, U, A, y, slist, m_list, y_preds):

        atts = get_attempts_from_db(psi)
        fatts = atts[atts["timestamp"] < ts]

        cats_visit_ct = Counter()
        cats_succ_ct = Counter()
        # dbvisited_pids =[]

        correct_qids = set(fatts[fatts["correct"] == True]["question_id"])
        visited_qids = set(fatts["question_id"])
        visited_pids = set()
        for qid in visited_qids:
            pid = qid.split("|")[0]
            if pid in all_page_ids:
                # if qid not in visited_qids:
                # if qid in fatts[fatts["correct"]==True]["question_id"]:
                cat = "TODO" # cat_page_lookup[pid]
                cats_visit_ct[cat] += 1
                if qid in correct_qids:
                    cats_succ_ct[cat] += 1
                # if pid not in dbvisited_pids:
                #     dbvisited_pids.append(pid)
                # visited_qids.append(qid)
                visited_pids.add(pid)

        maxlab="-"
        max_ixs_raw = list(reversed(list(y_predlist.argsort())))
        for mix in max_ixs_raw:
            mpid = pid_override[mix]
            if mpid not in visited_pids:
                maxlab = mpid
                break
        assert type(maxlab) is str

        natts = fatts.shape[0]
        nsucc = len(set(fatts[fatts["correct"]==True]["question_id"]))
        ndist = len(set(fatts["question_id"]))

        print(";;;;")
        # print(sorted(dbvisited_pids))
        print(sorted(visited_pids))

        # print(len(dbvisited_pids), len(visited_pids))
        # print(set(dbvisited_pids).symmetric_difference(set(visited_pids)))
        # assert sorted(dbvisited_pids) == sorted(visited_pids)

        visited_pids = "\n".join(map(str, visited_pids))

        assigned = []
        print("len(a) is {}",len(a))
        for ix,el in enumerate(a):
            if el > 0:
                print("access a at",ix)
                label = pid_override[ix]
                page = label.split("|")[0]
                if page not in assigned:
                    assigned.append(page)
                    print("ass page:",page)
        if len(assigned) > 0:
            assigned = "\n".join(map(str, assigned))
        else:
            assigned = "-"

        big5 = cats_succ_ct.most_common(20)
        if len(big5) == 0:
            big5 = "-"
        else:
            temp = []
            for cnt,succ in big5:
                v = cats_visit_ct[cnt]
                temp.append("{}: {} ({})".format(cnt,succ,v))
            big5 = "\n".join(temp)

        last5 = list(pandas.unique(fatts["question_id"])[-10:])
        temp5 = []
        for n in last5:
            if n in q_names_df.index:
                tit = q_names_df.loc[n, "title"]
                if str(tit)!="nan":
                    temp5.append("{} ({})".format(tit, n))
                else:
                    temp5.append(n)
            else:
                temp5.append("UNK")
        last5 = temp5
        last5 = "\n".join(map(str,last5))

        del fatts
        # if len(qh) > 0:
        #     ql, tl = zip(*qh)
        #     last5 = [q for q in numpy.unique(ql)[-5:]]
        #     last5 = "\n".join(last5)
        #     last5 = '{}'.format(last5)  # wrap in quotes
        # else:
        #     last5 = []



        print(s)
        for it in [psi, int(10 * s[0]) / 10.0, "{:.1f}".format(months_on), str(ndist)+" ("+str(natts)+")", nsucc, assigned, visited_pids, "_", last5, maxlab]:
            print(it)

        c=1
        for cv in [psi, int(10 * s[0]) / 10.0, "{:.1f}".format(months_on), str(ndist)+" ("+str(natts)+")", nsucc, assigned, visited_pids, big5, last5, str(mix)+":"+maxlab]:
            if cv == []:
                cv = "-"
            elif len(str(cv).split("\n")[0])>col_widths[c-1]:
                col_widths[c-1] = len(str(cv))
            ws.cell(r,c,cv)
            c += 1
        r += 1

    # for ci, cw in enumerate(col_widths):
    #     ws.column_dimensions[get_column_letter(ci + 1)].width = cw
    #
    # for ri, rh in enumerate(row_heights):
    #     ws.row_dimensions[ri+2].height = rh

    for col in ws.columns:
        max_length = 0
        column = col[0].column  # Get the column name
        for cell in col:
            cell.alignment = Alignment(horizontal="center", vertical="top")
            try:  # Necessary to avoid error on empty cells
                this_max = max([len(s) for s in str(cell.value).split('\n')])
                if this_max > max_length:
                    max_length = this_max
            except:
                pass
        adjusted_width = max_length * 1.2
        ws.column_dimensions[column].width = adjusted_width

    for row in ws.rows:
        max_height = 0
        rowname = row[0].row  # Get the column name
        for cell in row:
            try:  # Necessary to avoid error on empty cells
                cell_h = len(str(cell.value).split('\n'))
                print("for row {} cell value is {} at height {}".format(rowname, cell.value, cell_h))
                if cell_h > max_height:
                    # print("{} super {}, replaceing".format(cell_h, max_height))
                    max_height = cell_h
            except:
                pass
        adjusted_height = max_height * 11.5 # convert to points??
        ws.row_dimensions[rowname].height = adjusted_height
        
    wb.save('./report_cards/{:.1f}_{}_{}.xlsx'.format(months_av, gr_id, aid))

def create_student_scorecards(aug, model, all_qids=None, all_page_ids=None, pid_override=None):
    names_df = get_q_names()
    names_df.index = names_df["question_id"]

    aid_list, s_list, x_list, u_list, a_list, y_list, psi_list, hexes_to_try_list, hexes_tried_list, s_raw_list, gr_id_list, ts_list, maxdop = aug
    
    lookup = {}
    ts_grid_lookup = {}
    for aid,s,x,u,a,y,psi,grid,ts in zip(aid_list, s_list, x_list, u_list, a_list, y_list, psi_list, gr_id_list, ts_list):
        if aid not in lookup:
            lookup[aid] = ([],[],[],[],[],[])
            ts_grid_lookup[aid] = (ts,grid)
        sl,xl,ul,al,yl,psil = lookup[aid]
#         sr.append(s_raw)
        sl.append(s)
        xl.append(x)
        ul.append(u)
        al.append(a)
        yl.append(y)
        psil.append(psi)
        lookup[aid] = (sl,xl,ul,al,yl,psil)

    lkk = list(lookup.keys())

    for aid in lkk:
        m_list = []
        s_list = []
        x_list = []
        ts, gr_id = ts_grid_lookup[aid]
        sl, xl, ul, al, yl, psil = lookup[aid]
        predictions = []
        if(len(sl)==0):
            print("no students in aid:",aid)
            continue
        
        for s,x,u,psi in zip(sl,xl,ul,psil):
#             s_list.append(s)
#             x_list.append(x)
            # s_raw_list.append(s_raw)
            m_list.append(s[1])
            print("{}/{} done".format(aid,psi))

        if len(m_list)==0:
            print("skipping (no students) aid:",aid)
            continue
        print("m_list len:", len(m_list))
               
        s_arr = numpy.array(sl)
        x_arr = numpy.array(xl)
        u_arr = numpy.array(ul)
        a_arr = numpy.array(al)

        print(s_arr.shape, x_arr.shape, u_arr.shape, a_arr.shape)
        predictions = model.predict([s_arr,x_arr,u_arr,a_arr])
        save_class_report_card(ts, aid, gr_id, sl, xl, ul, al, yl, m_list, predictions, psil, names_df, all_qids, all_page_ids=all_page_ids, pid_override=pid_override)

    with open("a_ids.txt", "w+") as f:
        f.write("({})\n".format(len(aid_list)))
        f.writelines([str(a)+"\n" for a in sorted(aid_list)])
        f.write("\n")

