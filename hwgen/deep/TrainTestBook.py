import gc
import math
import os
import pickle
import zlib
from collections import Counter, defaultdict
from statistics import mean

import numpy
import pandas
import pylab
from keras import Input, Model
from keras.callbacks import EarlyStopping
from keras.layers import Dense, Dropout, concatenate, multiply
from keras.models import load_model
from keras.optimizers import Adam
from keras.utils import plot_model
from matplotlib import pyplot as plt, pyplot
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from sklearn.externals import joblib
from sklearn.linear_model import LogisticRegression, SGDClassifier
from sklearn.manifold import TSNE
from sklearn.neural_network import MLPClassifier
from sklearn.preprocessing import StandardScaler, LabelBinarizer
from sklearn.tree import DecisionTreeClassifier
from sklearn.utils import compute_class_weight

from hwgen.common import init_objects, get_meta_data, get_all_assignments, get_student_list, make_gb_question_map, \
    get_user_data, get_q_names
from hwgen.concept_extract import page_to_concept_map
from hwgen.hwgengen2 import hwgengen2, build_dob_cache
from hwgen.profiler import get_attempts_from_db

import openpyxl

base = "../../../isaac_data_files/"

n_users = -1
cats, cat_lookup, all_qids, users, diffs, levels, cat_ixs, cat_page_lookup, lev_page_lookup, all_page_ids = init_objects(
    n_users)

hwdf = get_meta_data()
concepts_all = set()
hwdf.index = hwdf["question_id"]
hwdf["related_concepts"] = hwdf["related_concepts"].map(str)
for concepts_raw in hwdf["related_concepts"]:
    print(concepts_raw)
    if concepts_raw != "nan":
        concepts = eval(concepts_raw)
        if concepts is not None:
            concepts_all.update(concepts)
concepts_all = list(concepts_all)

asst_fname = base + "assignments.pkl"
con_page_lookup = page_to_concept_map()


def cluster_and_print(assts):
    xygen = hwgengen2(assts, batch_size=-1, FRESSSH=False)  # make generator object
    for S, X, U, y, ai, awgt in xygen:
        y_labs = numpy.array(y)
        if (X == []):
            continue
        S = numpy.array(S)  # strings (labels)
        X = numpy.array(X)  # floats (fade)
        U = numpy.array(U)  # signed ints (-1,0,1)
        print("S", S.shape)
        print("X", X.shape)
        print("U", U.shape)
        assert y_labs.shape[1] == 1  # each line shd have just one hex assignment

        # c_labs = [concept_map[label[0]] for label in y_labs]
        # c = clb.transform(c_labs)
        # print(c_labs[0:10])
        # print(c[0:10])

        # try:
        #     y = ylb.transform(y_labs) / awgt
        # except:
        #     y = ylb.fit_transform(y_labs) / awgt
        # print(y_labs)
        # input(y)
        # input(awgt)
        # assert numpy.sum(y[0]) == 1 # Each line should be one-hot
        n = 5000
        lab_set = list(numpy.unique(y_labs))
        colors = numpy.array([lab_set.index(l) for l in y_labs])[0:n]

        # calc_entropies(X,y_labs)
        # exit()

        # pca = PCA(n_components=2)
        tsne = TSNE(n_components=2)
        # converted = pca.fit_transform(X) # convert experience matrix to points
        converted = tsne.fit_transform(X[0:n])

        plt.scatter(x=converted[:, 0], y=converted[:, 1], c=colors, cmap=pylab.cm.cool)
        plt.show()


def calc_entropies(X, y):
    d = defaultdict(list)
    for x, lab in zip(X, y):
        d[str(lab)].append(x)
    for l in d:
        # print("calc for {}, len {}".format(l,len(d[l])))
        H = entropy(d[l])
        print("{}\t{}\t{}".format(l, H, len(d[l])))


def entropy(lizt):
    "Calculates the Shannon entropy of a list"
    # get probability of chars in string
    lizt = [tuple(e) for e in lizt]
    prob = [float(lizt.count(entry)) / len(lizt) for entry in lizt]
    # calculate the entropy
    entropy = - sum([p * math.log(p) / math.log(2.0) for p in prob])
    return entropy


def make_phybook_model(n_S, n_X, n_U, n_A, n_P):

    # this is our input placeholder
    input_S = Input(shape=(n_S,), name="s_input")
    inner_S = Dense(10, activation="relu")(input_S)
    inner_S = Dropout(.2)(inner_S)
    inner_S = Dense(10, activation="relu")(inner_S)
    inner_S = Dropout(.2)(inner_S)
    # inner_S = Dense(10, activation="relu")(inner_S)
    # inner_S = Dropout(.2)(inner_S)

    w=300
    #w100: 17% to beat
    do=.2
    # input_X = Input(shape=(n_X,), name="x_input")
    # inner_X = Dense(w, activation="relu")(input_X)
    # inner_X = Dropout(.2)(inner_X)
    # inner_X = Dense(300, activation="relu")(inner_X)
    # inner_X = Dropout(.2)(inner_X)

    input_U = Input(shape=(n_U,), name="u_input")
    inner_U = Dense(w, activation="relu")(input_U)
    inner_U = Dropout(.2)(inner_U)
    inner_U = Dense(w, activation="relu")(inner_U)
    inner_U = Dropout(.2)(inner_U)
    # inner_U = Dense(w, activation="relu")(inner_U)
    # inner_U = Dropout(.2)(inner_U)

    # input_A = Input(shape=(n_A,), name="a_input")
    # inner_A = Dense(300, activation="relu")(input_A)
    # inner_A = Dropout(.2)(inner_A)
    # inner_A = Dense(n_P, activation="sigmoid")(inner_A)
    # inner_A = Dropout(.2)(inner_A)

    # mastery = concatenate([inner_X, inner_U])
    # mastery = Dense(300, activation="relu")(mastery)
    # mastery = Dropout(.2)(mastery)
    # mastery = Dense(300, activation="relu")(mastery)
    # mastery = Dropout(.2)(mastery)


    hidden = concatenate([inner_U, inner_S])
    # hidden = Dense(300, activation="relu")(hidden)
    # hidden = Dropout(.2)(hidden)

    # tabu = Dense(300, activation="relu")(input_A)
    # tabu = Dense(n_P, activation="sigmoid")(tabu)
    hidden = Dense(w, activation='relu')(hidden)
    hidden = Dropout(.2)(hidden)
    hidden = Dense(w, activation='relu')(hidden)
    hidden = Dropout(.2)(hidden)
    # hidden = Dense(w, activation='relu')(hidden)
    # hidden = Dropout(.2)(hidden)
    next_pg = Dense(n_P, activation='softmax')(hidden)
    # next_pg = multiply([next_pg, inner_A])
    # next_pg = Dense(n_P, activation='softmax', name="next_pg")(next_pg)

    o = Adam()

    m = Model(inputs=[input_S, input_U], outputs=next_pg)
    m.compile(optimizer=o, loss="categorical_crossentropy", metrics=["categorical_accuracy", "top_k_categorical_accuracy"])
    plot_model(m, to_file='hwgen_model.png')
    m.summary()
    input(",,,")
    return m


gb_qmap = make_gb_question_map()

numpy.set_printoptions(threshold=numpy.nan)


def augment_data(tr, sxua):
    concept_list = list(set().union(*concept_map.values()))
    print(concept_list)

    # yship = []
    # qlist = pid_override
    # print(qlist)
    # print("investigate this mofo")
    # exit()

    # hex_counter = Counter()
    # tot = 0

    # last_ts = None
    # for i, ass in enumerate(tr.iterrows()):
    #     ass_id = ass[1]["gameboard_id"]
    #     gb_id = ass[1]["gameboard_id"]
    #     gr_id = ass[1]["group_id"]
    #     ts = ass[1]["creation_date"]
    #     hex_acc = []
    #     if last_ts is not None and ((ts-last_ts).days==0):
    #         print("skipping same-day assignment")
            # hexagons = [gb_qmap[gb_id][0]]
            # for hx in hexagons:
            #     if hx not in hex_acc:
            #         hex_acc.append(hx)
            # continue # do not add same-day assignments

        # last_ts = ts
        # last_ass_id = ass_id

        # hexagons = [gb_qmap[gb_id][0]]
        # students = get_student_list(gr_id)
        # for psi in students:
        #     for hx in hexagons:
        #         if hx not in pid_override:
        #             print(hx, " not in qlist")
        #             pid_override.append(hx)
        #         yship.append(hx)
        #         hex_counter[hx] += 1
        #         tot += 1

    # yship = list(concept_map.keys()) +yship
    # ylb = LabelBinarizer()  # (classes=qlist)
    # qlist = numpy.unique(yship)
    # ylb.fit(qlist)
    # ylb.classes_ = yship  # start by fitting the binariser to the shortlist of book qns

    # for hx in hex_counter.most_common():
    #     print(hx[0], hx[1])
    # print(tot)

    # print(qlist)
    # print(ylb.classes_)
    # assert len(list(qlist)) == len(list(ylb.classes_))
    # assert list(qlist) == list(ylb.classes_)

    # weights = {}
    # class_wgt = compute_class_weight('balanced', ylb.classes_, yship)
    # for clix, (cls, wgt) in enumerate(zip(ylb.classes_, class_wgt)):
    #     print(clix, cls, wgt)
    #     weights[clix] = wgt

    group_ids = pandas.unique(tr["group_id"])

    aid_list = []
    s_list = []
    x_list = []
    u_list = []
    a_list = []
    y_list = []

    fout = open("tr_summ.csv","w")

    for gr_id in group_ids:
        gr_ass = tr[tr["group_id"] == gr_id]
        last_ts = None
        for row in gr_ass.iterrows():
            aid = row[1]["id"]
            ts = row[1]["creation_date"]
            gr_id = row[1]["group_id"]
            gb_id = row[1]["gameboard_id"]
            student_ids = list(get_student_list(gr_id)["user_id"])
            hexes= list(gb_qmap[gb_id])
            if last_ts is not None and ((ts-last_ts).days==0):
                print("skipping same-day assignment")
                continue # do not add same-day assignments
            last_ts = ts

            for psi in student_ids:
                S,X,U,A = pickle.loads(zlib.decompress(sxua[psi][ts]))
                if S[0]<10: #i.e. if student has no valid age
                    continue
                if S[1]==0: #no time in platform
                    continue

                hexes_tried = []
                hexes_to_try = []
                # if len(hexes)==1:
                #     hexes_to_try = hexes
                # else:
                # for ix, el in enumerate(X):
                #     if el > 0:
                #         page = all_qids[ix].split("|")[0]
                #         if page not in hexes_tried:
                #             hexes_tried.append(page)

                # for hx in hexes:
                #     if hx not in hexes_tried:
                #             hexes_to_try.append(hx)

                y_true = numpy.zeros(len(pid_override))  # numpy.zeros(len(all_page_ids))
                # for hx in hexes_to_try:

                for hx in hexes:
                    hxix = pid_override.index(hx)
                    if X[hxix]==0:
                        hexes_to_try.append(hx)

                if hexes_to_try==[]:
                    print("no hexes to try")
                    continue

                # decay = 0.5
                # w = 1.0
                # for hx in sorted(hexes_to_try):
                #     hxix = pid_override.index(hx)
                #     y_true[hxix] = 1 #/ len(hexes_to_try)
                #     if len(hexes_to_try)>1:
                #         print("trying", hx,w)
                #         input("")
                    # w = w * decay
                # y_true = y_true / y_true.sum()
                hxix = pid_override.index(sorted(hexes_to_try)[0])
                y_true[hxix] = 1.0

                    # else:
                    #     hexes_tried.append(hx)

                # hexes_tried = []
                # for i,el in enumerate(X):
                #     if el>0:
                #         pid = all_qids[i].split("|")[0]
                #         if pid not in hexes_tried:
                #             hexes_tried.append(pid)


                # print("hexes tried: {}".format(hexes_tried))
                print("hexes t try: {}".format(hexes_to_try))
                print("hexes      : {}".format(hexes))
                # print(numpy.sum(A))
                # print([all_page_ids[hx] for hx,el in enumerate(A) if el==1])
                aid_list.append(aid)
                # hex_list.append(hexes_to_try)
                s_list.append(S)
                # x_list.append(numpy.concatenate((X,U,A)))
                x_list.append(X)
                u_list.append(U)
                a_list.append(A)
                y_list.append(y_true)

                fout.write("{},{},{},{},{},{},{}\n".format(ts,psi,",".join(map(str,S)), X.sum(), numpy.sum(X>0), numpy.sum(U), " ".join(hexes_to_try)))
    fout.close()
    # exit()
    # input("nibit")
    gc.collect()
    s_list = numpy.array(s_list)
    x_list = numpy.array(x_list, dtype=numpy.int16)
    u_list = numpy.array(u_list, dtype=numpy.int8)
    a_list = numpy.array(a_list, dtype=numpy.int8)
    y_list = numpy.array(y_list, dtype=numpy.int8)
    return aid_list, s_list, x_list, u_list, a_list, y_list

def train_deep_model(tr, sxua, n_macroepochs=100, n_epochs=10, use_linear=False):
    model = None
    aid_list, s_list, x_list, u_list, a_list, y_list = augment_data(tr, sxua)

    SCALE = True
    sc = StandardScaler()
    if SCALE:
        # print(x_list.shape)
        lenX = s_list.shape[0]
        # for ix,x_el in enumerate(x_list):
        #     x_list[ix] = sc.transform(x_list[ix].reshape(1,-1))

        start = 0
        gap = 5000
        while(start<lenX):
            end = min(start+gap, lenX)
            print("fitting scaler",start,end)
            sc.partial_fit(s_list[start:end,:])
            start += gap

        start = 0
        while(start<lenX):
            end = min(start+gap, lenX)
            print("scaling",start,end)
            s_list[start:end,:] = sc.transform(s_list[start:end,:])
            start += gap

    print(s_list.shape, x_list.shape)
    # x_list = numpy.concatenate((s_list, x_list), axis=1)
    print(x_list.shape)
    gc.collect()

    #OK, we now have the four student profile vectors, and the true y vector, so we can fit the model
    if model is None:
        print("making model")
        S,X,U,A = s_list[0], x_list[0], u_list[0], a_list[0]
        print(S.shape, X.shape, U.shape, A.shape, y_list.shape)
        model = make_phybook_model(S.shape[0], X.shape[0], U.shape[0], A.shape[0], y_list.shape[1])
        print("model made")
        es = EarlyStopping(monitor='top_k_categorical_accuracy', patience=1, verbose=0, mode='auto')
        # es = EarlyStopping(monitor='categorical_accuracy', patience=0, verbose=0, mode='auto')
        history = model.fit([s_list, u_list], y_list, verbose=1, epochs=100, callbacks=[es], shuffle=True, batch_size=32)
        pyplot.plot(history.history['categorical_accuracy'])
        # pyplot.plot(history.history['binary_crossentropy'])
        # pyplot.plot(history.history['categorical_crossentropy'])
        pyplot.plot(history.history['top_k_categorical_accuracy'])
        pyplot.show()
        scores = model.evaluate([s_list,u_list],y_list)
        print(scores)

    return model, sc  # , sscaler, levscaler, volscaler


# def get_top_k_hot(raw, k): # TODO clean up
#     args = list(reversed(raw.argsort()))[0:k]
#     print(args)
#     k = numpy.zeros(len(raw))
#     k[args] = 1.0
#     return k, args


def evaluate_hits_against_asst(ailist, y, max_y, y_preds, ylb):
    '''This method should take an N~question assignment, compare it against the top N question candidates from the hwgen
    system and return a score between [0..1]'''
    true_qs = set(y)
    N = len(true_qs)
    # print("N",N)

    agg_cnt = Counter()
    for row in y_preds:
        args = numpy.argsort(row)
        topN = list(reversed(args))[0:N]
        for t in topN:
            # print(t)
            tlab = ylb.classes_[t]
            # print(tlab)
            agg_cnt[tlab] += 1
    print(agg_cnt.most_common(N))

    agg_pred = y_preds.sum(axis=0)
    print(agg_pred.shape)
    print(y_preds.shape)

    assert agg_pred.shape[0] == y_preds.shape[1]  # shd have same num of columns as y_preds
    sortargs = numpy.argsort(agg_pred)
    sortargs = list(reversed(sortargs))
    maxN = sortargs[0:N]  # get the top N
    pred_qs = [ylb.classes_[ix] for ix in maxN]
    print(true_qs, " vs ", pred_qs)
    score = len(true_qs.intersection(pred_qs)) / N
    print("score = {}".format(score))
    return score


def get_top_subjs(x, n):
    top_topics = Counter()
    for ix in range(len(x)):
        v = x[ix]
        if v != 0:
            qid = all_qids[ix]
            # print(qid)
            cat = cat_lookup[qid]
            top_topics[cat] += 1
    return top_topics.most_common(n)


def print_student_summary(ai, psi, s, x, ylb, clz, t, p, pr):
    top_subjs = get_top_subjs(x, ylb, 5)
    pid = None if clz is None else ylb.classes_[clz]
    xlevs = []
    for ix, el in enumerate(x):
        if el == 1:
            xlevs.append(levels[all_qids[ix]])

    print("{}:{}\t{}|\t{}\t{}\t{}\t({})\t{}\t{}\t{}\t{}".format(ai, psi, pr, t, p, pid, clz, numpy.sum(x), s,
                                                                numpy.mean(xlevs),
                                                                top_subjs))


def save_class_report_card(ts, aid, gr_id, S, X, U, A, y, y_preds, slist, q_names_df):

    N = y_preds.shape[0]
    print(N)
    sum_preds = numpy.sum(y_preds, axis=0)
    print("sum of sums", numpy.sum(sum_preds))
    sum_preds = sum_preds / N
    max_sum_ix = sum_preds.argmax()
    max_sum_prob = sum_preds.max()

    max_probs = y_preds.max(axis=1)
    max_ixs = y_preds.argmax(axis=1)
    max_labs = [pid_override[m] for m in max_ixs]

    vote_ct = Counter()
    for label in max_labs:
        vote_ct[label]+=1

    max_vote_lab = vote_ct.most_common(1)[0][0]
    max_sum_lab = pid_override[max_sum_ix]
    print("max sum lab =", max_sum_lab, max_sum_prob)
    print("votes counted:",vote_ct.most_common(5))
    print("most voted =", max_vote_lab)


    wb = openpyxl.Workbook()
    ws = wb.active

    fn_ai = str(aid)
    r = 1
    col_headers = ["student", "age", "months_on_isaac", "qns_tried", "successes", "prev_assignts", "hexes_attempted", "top_10_topics", "last_10_qns",
                   "ISAAC_SUGGESTS", "DIFF: (too easy 1..5 too hard)", "TOPIC:(bad 1..3 good)"]
    col_widths=[len(ch) for ch in col_headers]
    for c,cv in enumerate(col_headers):
        ws.cell(r,1+c,cv)

    ws.cell(2,2, ts)

    ws.cell(2,1, "Classroom sugg'n 1:")
    ws.cell(3,1, "Classroom sugg'n 2:")
    ws.cell(2,10, max_sum_lab)
    ws.cell(3,10, max_vote_lab)

    r=4
    months_on_list = []
    for s, x, u, a, t, psi, maxlab in zip(S, X, U, A, y, slist, max_labs):

        atts = get_attempts_from_db(psi)
        atts = atts[atts["timestamp"] < ts]

        pids =[]
        for qid in atts["question_id"]:
            pid = qid.split("|")[0]
            if pid not in pids:
                pids.append(pid)
        pids = "\n".join(map(str, pids))

        assigned = []
        for ix,el in enumerate(a):
            if el > 0:
                label = pid_override[ix]
                page = label.split("|")[0]
                if page not in assigned:
                    assigned.append(page)
        if len(assigned) > 0:
            assigned = "\n".join(map(str, assigned))

        big5 = get_top_subjs(x, 10)
        if len(big5) > 0:
            big5 = "\n".join(map(str, big5))

        natts = atts.shape[0]
        nsucc = atts[atts["correct"]==True].shape[0]

        last5 = list(pandas.unique(atts["question_id"])[-10:])
        temp5 = []
        for n in last5:
            tit = q_names_df.loc[n, "title"]
            if str(tit)!="nan":
                temp5.append("{} ({})".format(tit, n))
            else:
                temp5.append(n)
        last5 = temp5
        last5 = "\n".join(map(str,last5))


        # if len(qh) > 0:
        #     ql, tl = zip(*qh)
        #     last5 = [q for q in numpy.unique(ql)[-5:]]
        #     last5 = "\n".join(last5)
        #     last5 = '{}'.format(last5)  # wrap in quotes
        # else:
        #     last5 = []

        months_on = s[1] / 30.44
        months_on_list.append(months_on)

        c=1
        for cv in [psi, int(10 * s[0]) / 10.0, "{:.1f}".format(months_on), str(numpy.sum(x>0))+" ("+str(numpy.sum(x))+")", numpy.sum(u), [], pids, big5, last5, maxlab]:
            if cv == []:
                cv = "-"
            elif len(str(cv).split("\n")[0])>col_widths[c-1]:
                col_widths[c-1] = len(str(cv))
            ws.cell(r,c,cv)
            c += 1
        r += 1

    # for ci, cw in enumerate(col_widths):
    #     ws.column_dimensions[get_column_letter(ci + 1)].width = cw
    #
    # for ri, rh in enumerate(row_heights):
    #     ws.row_dimensions[ri+2].height = rh

    for col in ws.columns:
        max_length = 0
        column = col[0].column  # Get the column name
        for cell in col:
            cell.alignment = Alignment(horizontal="center", vertical="top")
            try:  # Necessary to avoid error on empty cells
                this_max = max([len(s) for s in str(cell.value).split('\n')])
                if this_max > max_length:
                    max_length = this_max
            except:
                pass
        adjusted_width = max_length * 1.2
        ws.column_dimensions[column].width = adjusted_width

    for row in ws.rows:
        max_height = 0
        rowname = row[0].row  # Get the column name
        for cell in row:
            try:  # Necessary to avoid error on empty cells
                cell_h = len(str(cell.value).split('\n'))
                print("for row {} cell value is {} at height {}".format(rowname, cell.value, cell_h))
                if cell_h > max_height:
                    print("{} super {}, replaceing".format(cell_h, max_height))
                    max_height = cell_h
            except:
                pass
        adjusted_height = max_height * 11.5 # convert to points??
        ws.row_dimensions[rowname].height = adjusted_height

    months_av = mean(months_on_list)
    wb.save('./report_cards/{:.1f}_{}_{}.xlsx'.format(months_av, gr_id, aid))

def create_student_scorecards(tt,sxua, model, sc):
    names_df = get_q_names()
    names_df.index = names_df["question_id"]
    cat_list = []
    ailist = []
    # all_page_ids = pid_override
    aids=[]
    for row in tt.iterrows():
        aid_list = []
        a_list = []
        x_list = []
        u_list = []
        y_list = []
        s_list = []
        hex_list = []
        psi_list = []

        print(row)
        aid = row[1]["id"]
        ts = row[1]["creation_date"]
        gr_id = row[1]["group_id"]
        gb_id = row[1]["gameboard_id"]
        student_ids = list(get_student_list(gr_id)["user_id"])
        print(student_ids)
        student_data = get_user_data(student_ids)
        hexes = list(gb_qmap[gb_id])
        print(hexes)

        # n-hot binarise the y vector here
        y_true = numpy.zeros(len(pid_override), dtype=numpy.int8)
        for hx in hexes:
            hxix = pid_override.index(hx)
            y_true[hxix] = 1.0

        aid_list.append(aid)
        incl_psis = []
        for psi in student_ids:
            S, X, U, A = pickle.loads(zlib.decompress(sxua[psi][ts]))
            if S[0]<10:
                print("s0 under 10")
                continue
            if S[1]==0:
                print("no time on plaform recorded")
                continue
            psi_list.append(psi)
            hex_list.append(hexes)
            y_list.append(y_true)
            # print(psi)
            # S,X,U,A = sxua[psi][ts]
            s_list.append(S)
            x_list.append(X)
            u_list.append(U)
            a_list.append(A)
            incl_psis.append(psi)
            print("student {} done".format(psi))

        if len(s_list)==0:
            continue

        s_arr = numpy.array(s_list)
        x_arr = numpy.array(x_list)
        u_arr = numpy.array(u_list)
        a_arr = numpy.array(a_list)

        predictions = model.predict([s_arr, u_arr])

        save_class_report_card(ts, aid, gr_id, s_list, x_list, u_list, a_list, y_list, predictions, incl_psis, names_df)
    with open("a_ids.txt", "w+") as f:
        f.write("({})\n".format(len(aid_list)))
        f.writelines([str(a)+"\n" for a in sorted(aids)])
        f.write("\n")

def evaluate_phybook_loss(tt,sxua, model, sc):
    aid_list, s_list, x_list, u_list, a_list, y_list = augment_data(tt, sxua)

    # hex_list = []
    # all_page_ids = pid_override
    # ailist = []
    for row in tt.iterrows():
        aid = row[1]["id"]
        # ts = row[1]["creation_date"]
        gr_id = row[1]["group_id"]
        gb_id = row[1]["gameboard_id"]
        student_ids = list(get_student_list(gr_id)["user_id"])
        print(student_ids)
        student_data = get_user_data(student_ids)
        hexes= list(gb_qmap[gb_id])
        print(hexes)

        for _ in student_ids:
            aid_list.append(aid)
            # hex_list.append(hexes)

    s_list = sc.transform(s_list)
    s_list = numpy.array(s_list)

    x_list = numpy.array(x_list)
    u_list = numpy.array(u_list)
    a_list = numpy.array(a_list)

    print(s_list.shape, x_list.shape, u_list.shape, a_list.shape)

    print("results")
    print(model.get_input_shape_at(0))
    predictions = model.predict([s_list, u_list])
    j_max = 0
    thresh_max = 0
    dir_hits_max = 0
    for j_thresh in [0.01, 0.025, .05, 0.075, .1,.2,0.3, 0.4, 0.5, 0.6, 0.7]:
    # for j_thresh in [0.4]:
        j_sum = 0
        # dir_sum = 0
        incl_sum = 0
        dir_hits = 0
        N = len(predictions)
        this_ai = None
        for ai, p, s,x,a,y in zip(aid_list, predictions, s_list,x_list,a_list,y_list):
            t = [pid_override[yix] for yix,yval in enumerate(y) if yval==1]
            if ai != this_ai:
                print("\n...new asst",ai)
                this_ai = ai
            phxs = []
            probs = []
            print("pshape",p.shape)
            maxpox = numpy.argmax(p)
            print(maxpox, len(pid_override))
            max_guess = pid_override[maxpox]
            phxs.append(max_guess)

            probs.append(p[maxpox])
            for ix, el in enumerate(p):
                if el>j_thresh and pid_override[ix] not in phxs:
                    phxs.append(pid_override[ix])
                    probs.append(p[ix])
            probs_shortlist = list(reversed(sorted(probs)))
            Z = list(reversed( [x for _, x in sorted(zip(probs, phxs))] ))
            # if Z:
            #     for t_el in t:
            #         if t_el in Z:#'direct hit'
            #             dir_sum += 1.0/len(t)
            print(t, Z)
            print(probs_shortlist)
            # print([all_page_ids[hx] for hx,el in enumerate(a) if el==1])
            if max_guess not in t:
                robot="BAD ROBOT"
            else:
                if max_guess == t[0]:
                    robot = "GREAT ROBOT"
                    dir_hits += 1
                else:
                    robot = "GOOD ROBOT"
            print("{} {}, XP={}".format(robot, sc.inverse_transform(s), numpy.sum(x)))
            t=set(t)
            phxs=set(phxs)
            if len(t.intersection(phxs)) > 0:
                incl_sum += 1
            j_sum  += len(t.intersection(phxs)) / len(t.union(phxs))
        j_score = j_sum/N
        # dir_score = dir_sum/N
        if dir_hits > dir_hits_max:
            j_max = j_score
            thresh_max = j_thresh
            dir_hits_max = dir_hits
            # dir_for_j_max = dir_score
        print("j_thresh =",j_thresh)
        print("Jaccard:", j_score)
        print("Incl:", incl_sum/N)
        print("D/H:", dir_hits/N)
        print("~ ~ ~ ~")
    print("max thresh/jacc:", thresh_max, j_max, dir_hits_max/N)
    print("num examples", N)

    #save_class_report_card(ailist,S,U,X,y,y_preds,awgt,slist,qlist,ylb)

    # print("max D/H:",dir_for_j_max)

def filter_assignments(assignments, book_only):
    # query = "select id, gameboard_id, group_id, owner_user_id, creation_date from assignments order by creation_date asc"
    assignments["include"] = True
    print(assignments.shape)
    map = make_gb_question_map()
    meta = get_meta_data()
    for ix in range(assignments.shape[0]):
        include = True
        gr_id = assignments.loc[ix, "group_id"]

        if book_only:
            gb_id = assignments.loc[ix, "gameboard_id"]
            hexes = map[gb_id]
            for hx in hexes:
                hx = hx.split("|")[0]
                if not (hx.startswith("ch_") or hx.startswith("ch-i")):
                    include = False
                    break

        if include:
            students = get_student_list([gr_id])
            if students.empty:
                include = False

        if include:
            include = False
            for psi in list(students["user_id"]):
                # print("checking",psi)
                atts = get_attempts_from_db(psi)
                if not atts.empty:
                    # print("OK")
                    include = True
                    break

        if not include:
            assignments.loc[ix, "include"] = False

    # assignments = assignments[assignments["include"]==True]
    print(assignments.shape)
    return assignments


concept_map = {}
topic_map = {}
concept_list = []
page_list = []
meta_df = pandas.DataFrame.from_csv(base + "book_question_meta.csv")
for thing in meta_df.iterrows():
    thing = thing[1]
    k = thing["URL:"].split("/")[-1]
    page_list.append(k)
    sft = "/".join((thing["Subject"], thing["Field"], thing["Topic"]))
    # concepts = thing["Related Concepts"].split(",")
    # concept_map[k] = concepts
    topic_map[k] = sft
    for c in concepts:
        if c not in concept_list:
            concept_list.append(c)

pid_override = list(topic_map.keys())

if __name__ == "__main__":
    # tracemalloc.start()
    print("Initialising deep learning HWGen....")

    os.nice(3)


    model = None
    # print("loading...")
    # with open(asst_fname, 'rb') as f:
    #     assignments = pickle.load(f)
    #
    # print("loaded {} assignments".format(len(assignments)))
    #

    USE_CACHED_ASSGTS = True
    SAVE_CACHED_ASSGTS = True
    cache_fname = base + "cached_assgts.csv"
    if USE_CACHED_ASSGTS:
        assignments = pandas.DataFrame.from_csv(cache_fname)
    else:
        assignments = get_all_assignments()
        assignments = filter_assignments(assignments, book_only=True)
        if SAVE_CACHED_ASSGTS:
            assignments.to_csv(cache_fname)
    # Now filter and split
    assignments = assignments[assignments["include"] == True]
    assignments["creation_date"] = pandas.to_datetime(assignments["creation_date"])

    BUILD_SXUA = False
    if BUILD_SXUA:
        print("building SXUA")
        SXUA = {}
        student_static = {}
        last_ts = {}
        last_hexes = {}
        print("build dob cache")
        try:
            dob_cache = joblib.load(base+"dob_cache")
        except:
            dob_cache= build_dob_cache(assignments)
            joblib.save(base+"dob_cache")
        print("done")

        group_ids = pandas.unique(assignments["group_id"])
        print(len(assignments))
        print(len(group_ids))
        print(group_ids[0:20])
        # exit()

        for gr_id in group_ids:
            gr_ass = assignments[assignments["group_id"]==gr_id]
            for row in gr_ass.iterrows():
        # for row in assignments.iterrows():
                aid = row[1]["id"]
                # print(row)
                ts = row[1]["creation_date"]
                # gr_id = row[1]["group_id"]
                gc.collect()
                gb_id = row[1]["gameboard_id"]
                student_ids = list(get_student_list(gr_id)["user_id"])
                # print(student_ids)
                student_data = get_user_data(student_ids)
                now_hexes= list(gb_qmap[gb_id])
                # print(now_hexes)
                # if 118651 not in student_ids:
                #     continue
                for psi in student_ids:
                    # if psi != 118651:
                    #     continue
                    # print(psi)
                    if psi not in SXUA:
                        S = numpy.zeros(6)
                        X = numpy.zeros(len(all_qids), dtype=numpy.int16)
                        U = numpy.zeros(len(all_qids), dtype=numpy.int8)
                        A = numpy.zeros(len(pid_override), dtype=numpy.int8)
                        SXUA[psi] = {}
                        print("+",psi, S, numpy.sum(X), numpy.sum(U), numpy.sum(A))
                        psi_data = student_data[student_data["id"]==psi]
                        rd = pandas.to_datetime(psi_data.iloc[0]["registration_date"])
                        # print(rd)
                        student_static[psi] = (rd,)
                        l_ts = pandas.to_datetime("1970-01-01 00:00:00")
                        l_hexes = []
                    else:
                        l_ts = last_ts[psi]
                        l_hexes = last_hexes[psi]
                        S,X,U,A = pickle.loads(zlib.decompress(SXUA[psi][l_ts]))
                    # S,X,U,A = copy(S),copy(X),copy(U),copy(A)
                    #make updates

                    # if psi ==118651:
                    #     print("birdskoeping")

                    attempts = get_attempts_from_db(psi)
                    attempts = attempts[attempts["timestamp"] < ts]
                    all_wins = list(attempts[(attempts["correct"] == True)]["question_id"])

                    recent_attempts = attempts[attempts["timestamp"]>=l_ts]
                    # qids = list(set(recent_attempts["question_id"]))
                    qids = list(set(recent_attempts["question_id"]))
                    recent_wins = list(recent_attempts[(recent_attempts["correct"] == True)]["question_id"])

                    for qid in qids:
                        try:
                            qix = all_qids.index(qid)
                            attct = (recent_attempts["question_id"] == qid).sum()
                            X[qix] += attct
                            if qid in recent_wins:
                                U[qix] = 1
                        except:
                            print("UNK Qn ", qid)
                            continue

                    print(l_hexes)
                    for hx in l_hexes:
                        hxix = pid_override.index(hx)
                        A[hxix] = 1

                    S[0] = (ts - dob_cache[psi]).days / 365.242 if dob_cache[psi] is not None else 0
                    # print(ts, l_ts)
                    day_delta = max(1, (ts-l_ts).seconds)/ 86400.0
                    att_delta = recent_attempts.shape[0]
                    all_atts = attempts.shape[0]
                    # print(day_delta, att_delta)
                    reg_date = student_static[psi][0]
                    # print(reg_date)
                    all_days = max(0, (ts - reg_date).days)
                    S[1] = all_days
                    S[2] = (att_delta/day_delta) #recent perseverence
                    S[3] = (len(recent_wins)/att_delta if att_delta else 0) # recent success rate
                    S[4] = (all_atts / all_days if all_days else 0) # all time perseverance
                    S[5] = (len(all_wins) / all_atts if all_atts else 0) # all time success rate

                    last_ts[psi] = ts
                    last_hexes[psi] = now_hexes
                    print("~",psi, S, numpy.sum(X), numpy.sum(U), numpy.sum(A))
                    SXUA[psi][ts] = zlib.compress(pickle.dumps((S,X,U,A)))
                    # if str(aid) in ["47150", "49320", "53792"]:
                    #     input(">> {}".format(aid))

        # f = open(base+"SXUA.pkl", 'wb')
        # pickle.dump(SXUA, f)
        # f.close()
        # print("*** *** *** SAVED")

        # print("compressing SXUA")
        # for st in SXUA:
        #     for tstamp in SXUA[st]:
        #         SXUA[st][tstamp] = zlib.compress(pickle.dumps(SXUA[st][tstamp]))
        f = open(base + "SXUA.comp.pkl", 'wb')
        pickle.dump(SXUA, f)
        f.close()
        print("compressed and SAVED")

    else:
        print("loading SXUA")
        f = open(base + "SXUA.comp.pkl", 'rb')
        SXUA = pickle.load(f)
        f.close()
    print("loaded")

    gc.collect()

    # assignments = assignments[assignments["owner_user_id"]==7062]
    # assignments = assignments[assignments["owner_user_id"]==6026]

    POST_FILTER=False
    if POST_FILTER:
        assignments["include"] = False
        for rix,row in enumerate(assignments.iterrows()):
            # print(row)
            aid = row[1]["id"]
            ts = row[1]["creation_date"]
            gr_id = row[1]["group_id"]
            gb_id = row[1]["gameboard_id"]
            student_ids = list(get_student_list(gr_id)["user_id"])
            # print(student_ids)
            # student_data = get_user_data(student_ids)
            hexes = list(gb_qmap[gb_id])
            # print(hexes)
            # if len(hexes)==1 and len(student_ids)<=20:
            if len(student_ids) <= 10:
                assignments.iloc[rix,6] = True

        print(assignments.shape[0])
        assignments = assignments[assignments["include"]==True]
        print(assignments.shape[0])

    COUNT_TEACHERS=True
    if COUNT_TEACHERS:
        ct = Counter()
        for t in numpy.unique(assignments["owner_user_id"]):
            t_assignments = assignments[assignments["owner_user_id"] == t]
            ct[t] = t_assignments.shape[0]
        print(ct.most_common(20))

    do_train = True
    do_testing = True
    create_scorecards = True

    t_train = None
    t_test = None


    asslimit = 20100
    assct = 0
    teacherct = 0
    for (t, tct) in list(ct.most_common(len(ct))):
        t_assignments = assignments[assignments["owner_user_id"] == t]
        this_ass = t_assignments.shape[0]
        if this_ass < 10:
            print("not enough assignments",t,tct,this_ass)
            del ct[t]
        else:
            teacherct += 1
            assct += this_ass
        if assct > asslimit:
            break
    print(teacherct, assct)

    teacherN = teacherct

    for (t, tct) in ct.most_common(teacherN):
        t_assignments = assignments[assignments["owner_user_id"] == t]
        print(t_assignments.shape[0], "new training assts")
        # sel_n = min(t_assignments.shape[0], assct // teacherN)
        # t_assignments = t_assignments.iloc[0:sel_n, :]

        nass = t_assignments.shape[0]
        if nass <2:
            this_split = 0
        else:
            # this_split = min(5,max(1,  nass//10))
            this_split = 1
        print("this split", this_split)
        # temp_t_tr = t_assignments.iloc[0:-(split//teacherN),:]
        # temp_t_tt = t_assignments.iloc[-(split//teacherN):, :]
        end = t_assignments.shape[0] - this_split
        temp_t_tr = t_assignments.iloc[0:end,:]
        temp_t_tt = t_assignments.iloc[end:, :]

        print("training dates:", temp_t_tr["creation_date"].min(), temp_t_tr["creation_date"].max())
        print("testing dates:", temp_t_tt["creation_date"].min(), temp_t_tt["creation_date"].max())
        if t_train is None:
            t_train = temp_t_tr
            t_test = temp_t_tt
            print("created t_train {} and t_test {}".format(len(t_train),len(t_test)))
        else:
            t_train = pandas.concat([t_train, temp_t_tr])
            t_test = pandas.concat([t_test, temp_t_tt])
            print("extended t_train {} and t_test {}".format(len(t_train),len(t_test)))
        # if len(t_train) + len(t_test) >= totass:
        #     print("exceeded totass")
        #     break
    tr = t_train
    tt = t_test

    gc.collect()
    print("Split complete!")
    print("{} {}".format(len(tt), len(tr)))
    # input(">")

    n_macroepochs =100
    n_epochs = 100

    if do_train:
        print("training")
        model, sc = train_deep_model(tr, SXUA, n_macroepochs, n_epochs)
        print("...deleted original X,y")
        model.save(base + 'hwg_model.hd5')
        joblib.dump(sc, base + 'hwg_mlb.pkl')
        # joblib.dump((sscaler,levscaler,volscaler), base + 'hwg_scaler.pkl')

    numpy.set_printoptions(precision=4)
    if do_testing:
        print("testing")
        if model is None:
            model = load_model(base + "hwg_model.hd5")
            sc = joblib.load(base + 'hwg_mlb.pkl')

        evaluate_phybook_loss(tt, SXUA, model, sc)  # , sscaler,levscaler,volscaler)
        # input("DEEP testing done")

    if create_scorecards:
        if model is None:
            model = load_model(base + "hwg_model.hd5")
            sc = joblib.load(base + 'hwg_mlb.pkl')
        create_student_scorecards(tt, SXUA, model, sc)
